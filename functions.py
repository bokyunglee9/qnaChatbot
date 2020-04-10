import openpyxl, nltk, string, math
from nltk import word_tokenize, pos_tag
from nltk.corpus import stopwords, wordnet
from nltk.stem.wordnet import WordNetLemmatizer
from textblob import TextBlob
from collections import Counter


def opening():
    print("\nThis is Python FAQ. \nChoose one category which you're going to ask.")
    user_question = input("Choose the number of category.\n1.Basic\n2.Objects\n3.Modules\n4.Data type\n5.GUI\nIf you want to quit chatbot, write 0(zero): ")
    program = True
    while program:
        if 0<= int(user_question) <=5:
            if int(user_question) == 1:
                category = 'basic'
                program = False
            elif int(user_question) == 2:
                category = 'objects'
                program = False
            elif int(user_question) == 3:
                category = 'modules'
                program = False
            elif int(user_question) == 4:
                category = 'data type'
                program = False
            elif int(user_question) == 5:
                category = 'gui'
                program = False
            elif int(user_question) == 0:
                category = 'Q'
                program = False
        else:
            print("Plz Try Again.")
            user_question = input(
                "Choose the number of category.\n1.Basic\n2.Objects\n3.Modules\n4.Data type\n5.GUI\nIf you want to quit chatbot, write 0(zero): ")
    return category

def readFAQ(category):
    Q = []
    A = []
    excel_document = openpyxl.load_workbook('python_qna.xlsx')
    sheet = excel_document.get_sheet_by_name(category)
    col_questions = 1  # the column that has questions
    col_answers = 2  # the column that has answers
    size = 0
    for row in sheet.rows:
        size += 1
    for i in range(size):
        Q.append(sheet.cell(row=i + 1, column=col_questions).value.lower())  # access a cell
        A.append(sheet.cell(row=i + 1, column=col_answers).value.lower())
    return Q,A

def printQ(Q):
    num = 1
    for question in Q:
        print("{}. {}".format(num, question))
        num += 1

def inputResponse():
    Response = []
    response = input("Write your question using the options above.: ")
    Response.append(response.lower())
    return Response

#Add question not in faq
def addFaq(new_question, category):
    wb = openpyxl.load_workbook('python_qna.xlsx')
    ws = wb.active
    ws = wb.get_sheet_by_name(category)
    col_questions = 1 # the column that has questions
    col_answers = 2 # the column that has answers
    size = 0
    for row in ws.rows:
        size = size + 1
    ws.cell(row=size+1, column=col_questions).value = new_question
    ws.cell(row=size+1, column= col_answers).value = 'in preparation'

    wb.save('python_qna.xlsx')
    wb.close()

##functions for data processing
#tokenization
def tokeni(Q):
    t= []
    for question in Q:
        tokens = word_tokenize(question)
        t.append(tokens)
    return t

#omitting stopwords
def omitStopword(t):
    tkns = []
    for sentence in t:
        sentence = [w for w in sentence if w not in stopwords.words('english') ]
        sentence = [w for w in sentence if w not in string.punctuation]
        tkns.append(sentence)
    return tkns

#pos-tagging
def posTag(tkns):
    tkns_pos = []
    for sentence in tkns:
        tokens = pos_tag(sentence)
        tkns_pos.append(tokens)
    return tkns_pos

#lemmatization
def lemmati(tkns_pos):
    wnl= nltk.WordNetLemmatizer()
    lem_pos =[]
    for sentence in tkns_pos:
        lem = []
        for word in sentence:
            if word[1].startswith('V'):
                w = wnl.lemmatize(word[0], pos='v')
                lem.append(w)
            elif word[1] == 'NNS':
                w = word[0][:-1]
                lem.append(w)
            elif word[1] == 'MD':
                w = wnl.lemmatize(word[0])
            else:
                w = wnl.lemmatize(word[0])
                lem.append(w)
        lem_pos.append(lem)
    return lem_pos

##data processing
#excel data
def processData(list):
    t = tokeni(list)
    tkns = omitStopword(t)
    tkns_pos = posTag(tkns)
    lem_pos = lemmati(tkns_pos)
    return lem_pos

#Cosine similarity
def get_cosine(vec1, vec2):
     intersection = set(vec1.keys()) & set(vec2.keys())
     numerator = sum([vec1[x] * vec2[x] for x in intersection])

     sum1 = sum([vec1[x]**2 for x in vec1.keys()])
     sum2 = sum([vec2[x]**2 for x in vec2.keys()])
     denominator = math.sqrt(sum1) * math.sqrt(sum2)

     if not denominator:
        return 0.0
     else:
        return float(numerator) / denominator

def cos_sim(user, q):
    sim= []
    vector1 = Counter(user)
    for sentence in q:
        vector2 = Counter(sentence)
        cosine = get_cosine(vector1, vector2)
        sim.append(cosine)
    return sim

def biggest_num(sim):
    num = 0
    for item in sim:
        if item > num:
            num = item
    if num > 0.33:
        return sim.index(num)
    else:
        return -1

def main(category):
    Q,A = readFAQ(category)
    printQ(Q)
    Response = inputResponse()
    lem_pos = processData(Q)
    user_lem_pos = processData(Response)
    user_lem_pos = user_lem_pos[0]
    sim = cos_sim(user_lem_pos, lem_pos)
    sim_num = biggest_num(sim)

    if sim_num == -1:
        print("\nI am sorry that I cannot answer you question now. I'll add your question then answer you next time.\n")
        addFaq(Response[0], category)

    else:
        my_answer = A[sim_num]
        print('\nAnswer:', my_answer)
        blob_answer_message = TextBlob(my_answer)
        translated = blob_answer_message.translate(to="ko")
        print('답변:', translated)
        terminate = False
        while not terminate:
            feedback = input("Is this the answer you wanted? (Y/N)").upper()
            if feedback == 'Y':
                print("I'm glad to answer you about python.")
                terminate = True
            elif feedback == 'N':
                print(
                    "I am sorry that I cannot answer your question now. I'll add your question then answer you next time.")
                addFaq(Response[0], category)
                terminate = True
            else:
                feedback = input("Is this the answer you wanted? (Y/N)").upper()
